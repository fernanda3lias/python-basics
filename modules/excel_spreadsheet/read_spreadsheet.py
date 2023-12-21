# openpyxl - reading data from an Excel spreadsheet (Workbook and Worksheet)
# doc: https://openpyxl.readthedocs.io/en/stable/

from pathlib import Path
from openpyxl import Workbook, load_workbook   
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Load excel file
workbook: Workbook = load_workbook(WORKBOOK_PATH)
# Name sheet
sheet_name = "My sheet"
# Select sheet
worksheet: Worksheet = workbook[sheet_name]

row: tuple[Cell]
for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')

        if cell.value == "Maria":
            worksheet.cell(cell.row, 2, 23)
    print()

#worksheet['B3'].value = 14

workbook.save(WORKBOOK_PATH)